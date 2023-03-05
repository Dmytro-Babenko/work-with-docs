from pathlib import Path
import re
from openpyxl import load_workbook, Workbook, worksheet
from win32com.client import Dispatch

GRAPH_SYMBOL = 'gr'

def get_info_from_tables(sheet: worksheet) -> dict[str:list]:
    info = {}
    for table in sheet.tables.values():    
        headers = table.column_names
        ref = re.sub(r'(\d):', lambda m: f'{int(m.group(1))+1}:', table.ref)
        info[table.name] = [{header: str(cell.value).replace('.', ',') if isinstance(cell.value, float) else cell.value
                            for header, cell in filter(lambda t: t[1].value, zip(headers, row))}
                            for row in sheet[ref] if row[0].value]
    return info

def get_info_from_defnames(book: Workbook):
    info = {}
    for name, defn_object in book.defined_names.items():
        destination = next(defn_object.destinations)
        sheet_name, coordinates = destination
        value = book[sheet_name][coordinates].value
        if isinstance(value, float):
            value = str(value).replace('.', ',')
        info[name] = value
    return info


def get_info_from_exel(exel_file: Path, sheet_name='info') -> dict[str:any]:
    '''
    Get information from the Exel sheet.
    Return dictionary
    '''
    wb = load_workbook(exel_file, data_only=True)
    info = {}
    ws = wb[sheet_name]

    info = get_info_from_tables(ws)
    info.update(get_info_from_defnames(wb))
    return info

    # for k, v in ws.iter_rows(values_only=True):
    #     if isinstance(v, float):
    #         v = str(v).replace('.', ',')
    #     info[k] = v
    # wb.close()
    # return info

def export_image(exel_file: Path, sheet_name) -> dict[str:Path]: 
    '''
    Save all charts in Exel sheet to the folder, with Exel_sheet name
    Return dictionary with images name and path
    '''
    graphs = []
    gr_folder = exel_file.parent.joinpath(GRAPH_SYMBOL)
    gr_folder.mkdir(exist_ok=True)
    app = Dispatch('Excel.Application')
    wb = app.Workbooks.Open(Filename=exel_file)
    app.DisplayAlerts = False

    i = 1
    gr_sheet = wb.Worksheets(sheet_name)
    for chartObject in gr_sheet.ChartObjects():
        # gr_name = f'{GRAPH_SYMBOL}{i}.png'
        gr_path = gr_folder.joinpath(f'{GRAPH_SYMBOL}{i}.png')
        graphs.append(gr_path)
        chartObject.Chart.Export(gr_path)
        i += 1
    wb.Close(SaveChanges=False, Filename=str(exel_file))

    return graphs

def first_sheet_name(exel_file: Path) -> str:
    '''Return name of the first sheet in Exel file'''
    wb = load_workbook(exel_file, read_only=True)
    sheet_name = wb.sheetnames[0]
    wb.close()
    return sheet_name

def is_sheet_exist(exel_file: Path, sheet_name: str) -> bool:
    '''Return True if sheet exists, else False'''
    wb = load_workbook(exel_file, read_only=True)
    if sheet_name in wb.sheetnames:
        wb.close()
        return True
    else:
        wb.close()
        return False
        
def existing_sheet(func):
    def inner(element, *args, **kwargs):
        output = func(element, *args, **kwargs)
        while not output:
            print('There no such sheet on exel file')
            output = func(element,  *args, **kwargs)
        return output
    return inner

@existing_sheet
def get_sheet_name(element, base, *args):
    sheet_name = input(f'\nWrite name of the {element}: ')
    if not is_sheet_exist(base, sheet_name):
        sheet_name = None
    return sheet_name

def rename_sheet(element, base, new_name, *args):
    '''Rename Exel sheet from old name to new'''
    book = load_workbook(base)
    old_name = get_sheet_name(element, base)
    book[old_name].title = new_name
    book.save(base)
    book.close()
    return new_name


def sheet_names(exel_file: Path) -> str:
    '''Return names list of the sheets in Exel file'''
    wb = load_workbook(exel_file, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names

