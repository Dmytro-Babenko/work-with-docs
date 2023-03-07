import win32com.client
from openpyxl import load_workbook

def copy_paste(path_exel, path_doc):
    excel = win32com.client.Dispatch('Excel.Application')
    word = win32com.client.Dispatch('Word.Application')

    wb = excel.Workbooks.Open(path_exel)
    ws = wb.Worksheets('gr')
    document = word.Documents.Open(path_doc)

    i = 0
    for chart in ws.ChartObjects():
        i+=1
        try:
            chart.Copy()
            bookmark = document.Bookmarks(f'gr{i}')
            bookmark.Range.Paste()
        except Exception as ex:
            print(type(ex))
            continue

    wb.Close(False)
    excel.Quit()
    document.SaveAs(r'D:\test\test.docx')
    document.Close()
    word.Quit()

path1 = r'D:\test\1.xlsx'
path2 = r'D:\test\gr.xlsx'

wb1 = load_workbook(path1, data_only=True)
wb2 = load_workbook(path2)
ws = wb1['gr']
table = ws.tables['gr']
ws2 = wb2['gr']
for i, row in enumerate(ws[table.ref]):
    for j, cell in enumerate(row):
        ws2.cell(row=i+1, column=j+1).value = cell.value

# data = [[cell.value for cell in row] for row in ws[table.ref]]
# print(data)
wb1.close()
wb2.save(path2)

# wb = load_workbook(path2)
# ws = wb.active
# height = len(data)
# width = len(data[0])
# for i in range(1, height+1):
#     for j in range(1, width+1):
#         ws.cell(row=i, column=j).value = data[i-1][j-1]
# wb.save(path2)
# path_doc = r'D:\test\копия.docx'

# copy_paste(path2, path_doc)