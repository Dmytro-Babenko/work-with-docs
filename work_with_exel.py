from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from statistics import mean
from openpyxl.styles import Font


data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = load_workbook('Grades.xlsx')
ws = wb.active

ws.title = 'Grades'

# make table header
x, *_ = data.values()
subjects = [*x.keys()]
header = ['Grades'] + subjects + ['Avr']
weight = len(header)
ws.append(header) 
for col in range(1, weight+1):
	ws.cell(1, col).font = Font(bold=True)

height = 1
for student, marks in data.items():
	height += 1
	marks = [*marks.values()]
	student_avg = mean(marks)
	# line = [student] + marks + [student_avg]
	line = [student] + marks + [f'=AVERAGE(B{height}:E{height})']
	ws.append(line)
    

height += 1
ws.cell(height, 1).value = 'avg'
for col in range(2, weight + 1):
	# marks_by_subjects = [ws.cell(x, col).value for x in range(2, height)]
	# sub_avg = mean(marks_by_subjects)
	c_l = get_column_letter(col)
	ws.cell(height, col).value = f'=AVERAGE({c_l}2:{c_l}{height-1})'

wb.save('Grades.xlsx')


