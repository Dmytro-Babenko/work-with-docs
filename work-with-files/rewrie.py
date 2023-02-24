from pathlib import Path
from work_with_exel import is_sheet_exist, first_sheet_name
from openpyxl import load_workbook

GRAPH_SYMBOL = 'gr'
TEMPLATE_SYMBOL = 'бланк'
RESULT_FOLDER_NAME = 'виконані'
EXTENTION = ['Sheet with charts']

MAIN_SETTINGS = {
    'base_folder': None,
    'Exel': None,
    'Word': None,
    'Sheet with information': None,
    'Sheet with charts': None,
    'Result file name': None,
    'Result folder': None
}

BASE = {
    'Exel': 'base_folder',
    'Word': 'base_folder',
    'Sheet with information': 'Exel',
    'Sheet with charts': 'Exel',
    'Result folder': 'base_folder'
}

def get_settings(settings: dict):

    def confirmation(func) -> bool:
        def inner(element, *args, **kwargs):
            value = settings[element]
            if value == None:
                return func(element, *args, **kwargs)
            if element in EXTENTION:
                return value
            confirmation = input(f'\n{element}: {value}\nIf you confirm print "yes", else - "no" and press Enter: ')
            while True:
                confirmation = confirmation.lower().strip()
                if confirmation == 'yes':
                    return value
                elif confirmation == 'no':
                    return func(element, *args, **kwargs)
                else:
                    confirmation = input('Sorry, write "yes" to confirm, or "no" in others cases: ')
        return inner

    def existing_path(func):
        def inner(element, *args, **kwargs):
            output = func(element, *args, **kwargs)
            while not output.exists():
                print('There no file with such path')
                output = func(element,  *args, **kwargs
                )
            return output
        return inner

    def existing_sheet(func):
        def inner(element, *args, **kwargs):
            output = func(element, *args, **kwargs)
            while not output:
                print('There no such sheet on exel file')
                output = func(element,  *args, **kwargs)
            return output
        return inner

    @confirmation
    @existing_path
    def get_path_by_name(element, base, *args):
        inp = input(f'\nWrite name of the {element}: ')
        path = base.joinpath(inp)
        return path

    @confirmation
    @existing_path
    def get_fullpath(element, *args):
        path = Path(input(f'\nWrite path of the {element}: '))
        return path

    @confirmation
    @existing_sheet
    def get_sheet_name(element, base, *args):
        sheet_name = input(f'\nWrite name of the {element}: ')
        if sheet_name not in load_workbook(base).sheetnames:
            sheet_name = None
        return sheet_name

    @confirmation
    def rename_sheet(element, base, new_name, *args):
        book = load_workbook(base)
        old_name = get_sheet_name(element, base)
        book[old_name].name = new_name
        book.save(str(base.absolute())) #разобраться
        book.close()
        return new_name

    @confirmation
    def get_name(element, *args):
        name = input(f'\n{element}: ')
        return name
        
    def accept(element, *args):
        return settings[element]

    user_settings_funcs = {
        'base_folder': get_fullpath,
        'Exel': get_path_by_name,
        'Word': get_path_by_name,
        'Sheet with information': get_sheet_name,
        'Sheet with charts': rename_sheet,
        'Result file name': get_name,
        'Result folder': accept
    }
        
    def get_defoult_doc(base: Path):
        pattern = f'*{TEMPLATE_SYMBOL}*.docx'
        output = next(base.glob(pattern))
        return output
    
    def get_defoult_exel(base:Path):
        pattern = f'*{TEMPLATE_SYMBOL}*.xlsx'
        output = next(base.glob(pattern))
        return output
    
    def get_defoult_infosheet(base:Path):
        return first_sheet_name(base)

    def get_defoult_chartsheet(base:Path):
        if is_sheet_exist(base, GRAPH_SYMBOL):
            return GRAPH_SYMBOL
        return None
    
    def get_defoult_resultfolder(base: Path):
        output = base.joinpath(RESULT_FOLDER_NAME)
        output.mkdir(exist_ok=True)
        return output
    
    def no_defoult_settings(base: Path):
        return None

    defoult_settings_funcs = {
        'base_folder': no_defoult_settings,
        'Exel': get_defoult_exel,
        'Word': get_defoult_doc,
        'Sheet with information': get_defoult_infosheet,
        'Sheet with charts': get_defoult_chartsheet,
        'Result file name': no_defoult_settings,
        'Result folder': get_defoult_resultfolder
    }
    
    def choose_func(element, hendler):
        return hendler[element]
    
    def choose_base(element):
        base = settings.get(BASE.get(element))
        return base
    
    for element in settings:
        base = choose_base(element)
        get_defoult_value = choose_func(element, defoult_settings_funcs)
        settings[element] = get_defoult_value(base)
        get_user_value = choose_func(element, user_settings_funcs)
        settings[element] = get_user_value(element, base, GRAPH_SYMBOL)
    return settings

get_settings(MAIN_SETTINGS)
print(MAIN_SETTINGS)



# path = Path(r'D:\test').glob('*бланк*.docx')
# print(next(path))
# for i in path:
#     print(i)

# Path(r'D:\test\1').mkdir(exist_ok=True)