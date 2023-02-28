from pathlib import Path
from openpyxl import load_workbook
from win32com.client import Dispatch

GRAPH_SYMBOL = 'gr'

def get_info_from_exel(exel_file: Path, sheet_name='info') -> dict[str:any]:
    '''
    Get information from the Exel sheet.
    Information shouuld be wtitten in first two column.
    Return dictionary (key - firsrt_column, value -second)
    '''
    wb = load_workbook(exel_file, data_only=True, read_only=True)
    info = {}
    ws = wb[sheet_name]
    for k, v in ws.iter_rows(values_only=True):
        if isinstance(v, float):
            v = str(v).replace('.', ',')
        info[k] = v
    wb.close()
    return info

def export_image(exel_file: Path, sheet_name) -> dict[str:Path]: 
    '''
    Save all charts in Exel sheet to the folder, with Exel_sheet name
    Return dictionary with images name and path
    '''
    graphs = {}
    gr_folder = exel_file.parent.joinpath(GRAPH_SYMBOL)
    gr_folder.mkdir(exist_ok=True)
    app = Dispatch('Excel.Application')
    wb = app.Workbooks.Open(Filename=exel_file)
    app.DisplayAlerts = False

    i = 1
    gr_sheet = wb.Worksheets(sheet_name)
    for chartObject in gr_sheet.ChartObjects():
        gr_name = f'{GRAPH_SYMBOL}{i}.png'
        gr_path = gr_folder.joinpath(gr_name)
        graphs[gr_name] = gr_path
        chartObject.Chart.Export(gr_path)
        i += 1
    wb.Close(SaveChanges=False, Filename=str(exel_file))

    return graphs

def first_sheet_name(exel_file: Path) -> str:
    '''Return name of the first sheet in Exel file'''
    wb = load_workbook(exel_file, read_only=True)
    sheet_name = wb.sheetnames[0]
    wb.close()
    return sheet_name

def is_sheet_exist(exel_file: Path, sheet_name: str) -> bool:
    '''Return True if sheet exists, else False'''
    wb = load_workbook(exel_file, read_only=True)
    if sheet_name in wb.sheetnames:
        wb.close()
        return True
    else:
        wb.close()
        return False
        
# def rename_sheet(exel_file: Path, old_sheet_name: str, new_sheet_name: str) -> None:
#     '''
#     Rename Exel sheet from old name to new. 
#     If sheet with this name doesnt exist, create sheet with new name.
#     '''
#     book = load_workbook(exel_file)
#     if old_sheet_name in book.sheetnames:
#         book[old_sheet_name].name = new_sheet_name
#     else:
#         book.create_sheet(new_sheet_name)
#     book.save(exel_file)
#     pass

def existing_sheet(func):
    def inner(element, *args, **kwargs):
        output = func(element, *args, **kwargs)
        while not output:
            print('There no such sheet on exel file')
            output = func(element,  *args, **kwargs)
        return output
    return inner

@existing_sheet
def get_sheet_name(element, base, *args):
    sheet_name = input(f'\nWrite name of the {element}: ')
    if not is_sheet_exist(base, sheet_name):
        sheet_name = None
    return sheet_name

def rename_sheet(element, base, new_name, *args):
    '''Rename Exel sheet from old name to new'''
    book = load_workbook(base)
    old_name = get_sheet_name(element, base)
    book[old_name].title = new_name
    book.save(base)
    book.close()
    return new_name




# wb = load_workbook(r'D:\PYHTON\GitHub\work-with-docs\work-with-files\1.xlsx', data_only=True, read_only=True)
# ws = wb['asda']
# for k, v in ws.iter_rows(values_only=True):
#     if isinstance(v, float):
#         print(v)
# wb.close()
