import re
from setting_class import Setting, DATA_MAIN_SETTINGS, GRAPH_SYMBOL
from openpyxl import worksheet, load_workbook, Workbook
from docxtpl import DocxTemplate, InlineImage
from clean_folder.sort import find_free_name
from win32com.client import Dispatch
from pathlib import Path

class ExelComplex():
    def __init__(self, main_path:Path, graph_path:Path, infosheet:str, graphsheet:str) -> None:
        self.main_path = main_path
        self.graph_path = graph_path
        self.infosheet = infosheet
        self.graphsheet = graphsheet
        pass

    def get_info_from_tables(self, sheet: worksheet) -> dict[str:list]:
        info = {}
        for table in sheet.tables.values():    
            headers = table.column_names
            ref = re.sub(r'(\d):', lambda m: f'{int(m.group(1))+1}:', table.ref)
            info[table.name] = [
                {header: str(round(cell.value, 3)).replace('.', ',') if isinstance(cell.value, float) else cell.value
                for header, cell in filter(lambda t: t[1].value, zip(headers, row))}
                for row in sheet[ref] if row[0].value
                ]
        return info

    def get_info_from_defnames(self, book: Workbook):
        info = {}
        for name, defn_object in book.defined_names.items():
            destination = next(defn_object.destinations)
            sheet_name, coordinates = destination
            value = book[sheet_name][coordinates].value
            if isinstance(value, float):
                value = str(round(value, 3)).replace('.', ',')
            info[name] = value
        return info


    def get_info(self) -> dict[str:any]:
        '''
        Get information from the Exel sheet.
        Return dictionary
        '''
        wb = load_workbook(self.main_path, data_only=True)
        ws = wb[self.infosheet]
        info = {}

        info = self.get_info_from_tables(ws)
        info.update(self.get_info_from_defnames(wb))
        wb.close()
        return info
    
    def export_image(self) -> dict[str:Path]: 
        '''
        Save all charts in Exel sheet to the folder, with Exel_sheet name
        Return dictionary with images name and path
        '''
        graphs = []
        gr_folder = self.path.parent.joinpath(GRAPH_SYMBOL)
        gr_folder.mkdir(exist_ok=True)
        app = Dispatch('Excel.Application')
        wb = app.Workbooks.Open(Filename=self.main_path)
        app.DisplayAlerts = False

        i = 1
        gr_sheet = wb.Worksheets(self.graphsheet)
        for chartObject in gr_sheet.ChartObjects():
            gr_path = gr_folder.joinpath(f'{GRAPH_SYMBOL}{i}.png')
            graphs.append(gr_path)
            chartObject.Chart.Export(gr_path)
            i += 1
        wb.Close(SaveChanges=False, Filename=str(self.main_path))
        app.Quit()
        return graphs
    
    def change_graph_file(self):
        main_wb = load_workbook(self.main_path, data_only=True)
        main_ws = main_wb[self.graphsheet]
        table = main_ws.tables[GRAPH_SYMBOL]

        graph_wb = load_workbook(self.graph_path)
        graph_ws = graph_wb.active
        # graph_ws.title = GRAPH_SYMBOL
        
        ref = re.sub(r'(\d):', lambda m: f'{int(m.group(1))+1}:', table.ref)
        for i, row in enumerate(main_ws[ref]):
            for j, cell in enumerate(row):
                graph_ws.cell(row=i+1, column=j+1).value = cell.value if cell.value != ' ' else None

        main_wb.close()
        graph_wb.save(self.graph_path)

    def copy_paste_charts(self, document_path: Path):
        excel = Dispatch('Excel.Application')
        word = Dispatch('Word.Application')

        try:
            wb = excel.Workbooks.Open(self.graph_path)
            ws = wb.Worksheets(1)
            d_path = str(document_path.absolute())
            document = word.Documents.Open(d_path)

            for i, chart in enumerate(ws.ChartObjects()):
                try:
                    chart.Copy()
                    bookmark = document.Bookmarks(f'{GRAPH_SYMBOL}{i+1}')
                    bookmark.Range.Paste()
                except Exception:
                    continue
        finally:
            wb.Close(False)
            excel.Quit()
            document.SaveAs(d_path)
            document.Close()
            word.Quit()


class MainSettings(Setting):
    def __init__(self, setting_name, data) -> None:
        super().__init__(setting_name, data)

    def exe_program(self):
        exel_path = self.data['exel'].get_value()
        doc_path = self.data['word'].get_value()
        exel_gr_path = self.data['exel with charts'].get_value()
        text_sheet = self.data['sheet with information'].get_value()
        graph_sheet = self.data['sheet with charts'].get_value()
        result_file_name = self.data['result file name'].get_value()
        result_folder = self.data['result folder'].get_value()

        doc = DocxTemplate(doc_path)
        exel = ExelComplex(exel_path, exel_gr_path, text_sheet, graph_sheet)
        info = exel.get_info()
        # graphs = exel.export_image()
        # graphs = {x.stem: InlineImage(doc, str(x)) for x in graphs}
        # info.update(graphs)
        doc_result_path = find_free_name(result_file_name, result_folder, doc_path.suffix)[1]
        doc.render(info)
        doc.save(doc_result_path)
        exel.change_graph_file()
        exel.copy_paste_charts(doc_result_path)
        pass


def main():
    main_settings = MainSettings('Main settings', DATA_MAIN_SETTINGS)
    main_root = main_settings.make_setting_root()
    pass_frame = main_settings.make_frame(main_root)
    main_settings.make_password_root(pass_frame)
    main_root.mainloop()

if __name__ == '__main__':
    main()