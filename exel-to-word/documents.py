import re
from setting_class import GRAPH_SYMBOL
from openpyxl import worksheet, load_workbook, Workbook
from win32com.client import Dispatch
from pathlib import Path

CONFIRM = 'Confirm'
RESET = 'Reset'
CONFIRM_MESSAGE = 'You realy confirm it?'
DONE_MESSAGE = 'Successfully completed'
ERROR_MESSAGE = 'Something go wrong'
FIELD_CONFIGURATION = {
    'borderwidth': 2, 
    'width': 100
}

PASS_VARIABLE = 'pass_var'
PASSWORD = 'bvv23015'
PASSWORD_BUTTON = 'Confirm password'
PASSWORD_ERROR = 'Wrong password'
PASS_CONFIGURATION = {
    'borderwidth': 2, 
    'width': 20
}

class ExelComplex():
    def __init__(self, main_path:Path, graph_path:Path, infosheet:str, graphsheet:str) -> None:
        self.main_path = main_path
        self.graph_path = graph_path
        self.infosheet = infosheet
        self.graphsheet = graphsheet
        self.app = None
        self.graphfile = None
        pass

    def __enter__(self):
        try:
            self.app = Dispatch('Excel.Application')
            self.graphfile = self.app.Workbooks.Open(self.graph_path)
            return self.graphfile
        except Exception as e:
            self.__exit__(type(e), e, e.__traceback__)
            raise e

    
    def __exit__(self, exception_type, exception_value, traceback):
        if self.graphfile:
            self.graphfile.Close(False)
        if self.app:
            self.app.Quit()

    def charts_check(func):
        def inner(self, *args):
            main_wb = load_workbook(self.main_path, data_only=True)
            main_ws = main_wb[self.graphsheet]
            if GRAPH_SYMBOL in main_ws.tables:
                main_wb.close
                func(self, *args)
            main_wb.close
            pass
        return inner

    def get_info_from_tables(self, sheet: worksheet) -> dict[str:list]:
        info = {}
        for table in sheet.tables.values():    
            headers = table.column_names
            ref = re.sub(r'(\d+):', lambda m: f'{int(m.group(1))+1}:', table.ref)
            info[table.name] = [
                {header: str(round(cell.value, 3)).replace('.', ',') if isinstance(cell.value, float) else cell.value
                for header, cell in filter(lambda t: t[1].value, zip(headers, row))}
                for row in sheet[ref] if row[0].value and str(row[0].value).strip()
                ]
        return info

    def get_info_from_defnames(self, book: Workbook):
        info = {}
        for name, defn_object in book.defined_names.items():
            destination = next(defn_object.destinations)
            sheet_name, coordinates = destination
            value = book[sheet_name][coordinates].value
            if isinstance(value, float):
                value = str(round(value, 3)).replace('.', ',')
            info[name] = value
        return info


    def get_info(self) -> dict[str:any]:
        '''
        Get information from the Exel sheet.
        Return dictionary
        '''
        wb = load_workbook(self.main_path, data_only=True)
        ws = wb[self.infosheet]
        info = {}

        info = self.get_info_from_tables(ws)
        info.update(self.get_info_from_defnames(wb))
        wb.close()
        return info
    
    def export_image(self) -> dict[str:Path]: 
        '''
        Save all charts in Exel sheet to the folder, with Exel_sheet name
        Return dictionary with images name and path
        '''
        graphs = []
        gr_folder = self.path.parent.joinpath(GRAPH_SYMBOL)
        gr_folder.mkdir(exist_ok=True)
        app = Dispatch('Excel.Application')
        wb = app.Workbooks.Open(Filename=self.main_path)
        app.DisplayAlerts = False

        i = 1
        gr_sheet = wb.Worksheets(self.graphsheet)
        for chartObject in gr_sheet.ChartObjects():
            gr_path = gr_folder.joinpath(f'{GRAPH_SYMBOL}{i}.png')
            graphs.append(gr_path)
            chartObject.Chart.Export(gr_path)
            i += 1
        wb.Close(SaveChanges=False, Filename=str(self.main_path))
        app.Quit()
        return graphs
    
    @charts_check
    def change_graph_file(self):
        main_wb = load_workbook(self.main_path, data_only=True)
        main_ws = main_wb[self.graphsheet]
        table = main_ws.tables[GRAPH_SYMBOL]

        graph_wb = load_workbook(self.graph_path)
        graph_ws = graph_wb.active
        
        ref = re.sub(r'(\d):', lambda m: f'{int(m.group(1))+1}:', table.ref)
        for i, row in enumerate(main_ws[ref]):
            for j, cell in enumerate(row):
                graph_ws.cell(row=i+1, column=j+1).value = cell.value if cell.value != ' ' else None

        main_wb.close()
        graph_wb.save(self.graph_path)
        pass

    @charts_check
    def copy_paste_charts(self, document_path: Path):
        with self as wb:
            with WordFile(document_path) as document:
                ws = wb.Worksheets(1)
                for i, chart in enumerate(ws.ChartObjects()):
                    try:
                        chart.Copy()
                        bookmark = document.Bookmarks(f'{GRAPH_SYMBOL}{i+1}')
                        bookmark.Range.Paste()
                    except Exception:
                        continue
        pass

class WordFile:
    def __init__(self, path) -> None:
        self.path = str(path)
        self.app = None
        self.file = None
        pass

    def __enter__(self):
        try:
            self.app = Dispatch('Word.Application')
            self.file = self.app.Documents.Open(self.path)
            return self.file
        except Exception as e:
            self.__exit__(type(e), e, e.__traceback__)
            raise e
    
    def __exit__(self, exception_type, exception_value, traceback):
        if self.file:
            self.file.Close(SaveChanges=True)
        if self.app:
            self.app.Quit()