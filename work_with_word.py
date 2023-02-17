from pathlib import Path
from docxtpl import DocxTemplate, InlineImage
from openpyxl import Workbook, load_workbook
from win32com.client import Dispatch
from docx.shared import Cm

def get_pathes(parent_folder=None|Path, is_folder = True, word='folder') -> Path:
    n_p = 'name' if parent_folder else 'path'
    inp = input(f'Write {n_p} of the {word}: ')
    path = parent_folder.joinpath(inp) if parent_folder else Path(inp)
    while True:
        if path.exists():
            break
        else:
            inp = input(f'Write {n_p} of the {word}: ')
            path = parent_folder.joinpath(inp) if parent_folder else Path(inp)
    return Path
    # else:



    #     path = input('There are no {word} on this path. Please write another or create it: ')
    #         folder = pathlib.Path(path)

    
    # exel_name = input('Write name of the Exel file: ')
    # exel_file = folder.joinpath(exel_name)
    # word_blank_name = input('Write name of the Word-blank: ')
    # word_blank = folder.joinpath(word_blank_name)



def get_info_from_exel(wb: Workbook) -> dict[str:any]:
    info = {}
    ws = wb['text']
    for k, v in ws.iter_rows(values_only=True):
        info[k] = v
    return info

def export_image(exel_file: Path, sheet_name='gr') -> dict[str:Path]:
    graphs = {}
    gr_folder = exel_file.parent.joinpath(sheet_name)
    if not gr_folder.exists():
        gr_folder.mkdir()
    app = Dispatch('Excel.Application')
    wb = app.Workbooks.Open(Filename=exel_file)
    app.DisplayAlerts = False

    i = 1
    gr_sheet = wb.Worksheets(sheet_name)
    for chartObject in gr_sheet.ChartObjects():
        gr_name = f'{sheet_name}{i}.png'
        gr_path = gr_folder.joinpath(gr_name)
        graphs[gr_name] = gr_path
        chartObject.Chart.Export(gr_path)
        i += 1
    wb.Close(SaveChanges=False, Filename=str(exel_file))
    
    return graphs
    
def create_image_template(doc: DocxTemplate, folder: Path) -> None:
    placeholders = {}
    i = 1
    for image in folder.iterdir():
        if image.suffix == '.png':
            placeholder = InlineImage(doc, str(image.absolute()), Cm(5))
            place = f'gr{i}'
            placeholders[place] = placeholder
            i += 1
    doc.render(placeholders)
    doc.save('test1.docx')
    pass
    

def main():
    


    file_exel = Path('D:\PYHTON\GitHub\work-with-docs\info.xlsm')
    wb = load_workbook(file_exel, data_only=True)
    doc = DocxTemplate('test.docx')
    graphs = export_image(file_exel)
    for plate, graph in graphs.items():
        doc.replace_pic(plate, graph)
    info = get_info_from_exel(wb)
    doc.render(info)
    doc.save('rename.docx')

if __name__ == '__main__':
    main()
# doc = DocxTemplate('test.docx')
# context = {'NAME':'Sven'}

# doc = DocxTemplate('test.docx')
# folder = Path('D:\PYHTON\GitHub\work-with-docs\Template')
# create_image_template(doc, folder)


