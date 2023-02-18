from pathlib import Path
from openpyxl import load_workbook
def get_path(word: str, parent_folder=None) -> Path:
    n_p = 'name' if parent_folder else 'path'
    inp = input(f'Write {n_p} of the {word}: ')
    if parent_folder:
        path = parent_folder.joinpath(inp)
    path = parent_folder.joinpath(inp) if parent_folder else Path(inp)
    while True:
        if path.exists():
            break
        else:
            inp = input(f'Write {n_p} of the {word}: ')
            path = parent_folder.joinpath(inp) if parent_folder else Path(inp)
    return path

def get_sheet_name(exel_path: Path, word: str) -> str:
    name = input(f'Write name of the exel sheet with {word}: ')
    exel_file = load_workbook(exel_path)
    while name not in exel_file.sheetnames:
        name = input(f'There no this sheet in the file. Write name of the exel sheet with {word}: ')
    return name



